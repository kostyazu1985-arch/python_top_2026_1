import requests
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================
# НАСТРОЙКИ ПАРСЕРА
# ============================================

# Базовый URL сайта
BASE_URL = "https://ria.ru"

# Заголовки для имитации браузера
# ВАЖНО: Без правильных заголовков сайт может блокировать запросы
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1"
}


# ============================================
# ФУНКЦИЯ ПАРСИНГА НОВОСТЕЙ
# ============================================

def parse_ria_news():
    """
    Парсит новости с главной страницы РИА Новости

    Возвращает:
        Список словарей с информацией о новостях
    """

    # Создаем пустой список для хранения новостей
    all_news = []

    try:
        # Выводим информацию о начале парсинга
        print("📰 Начинаем парсинг новостей с РИА Новости...")
        print(f"🌐 URL: {BASE_URL}")
        print("-" * 60)

        # Отправляем GET-запрос к сайту с заголовками
        # timeout=10 - максимальное время ожидания ответа (10 секунд)
        response = requests.get(BASE_URL, headers=headers, timeout=10)

        # Проверяем статус ответа
        # 200 - успешный запрос
        if response.status_code != 200:
            print(f"❌ Ошибка: сервер вернул статус {response.status_code}")
            return all_news

        print("✅ Страница успешно загружена!")

        # Создаем объект BeautifulSoup для парсинга HTML
        # 'lxml' - используем парсер lxml (быстрее, чем 'html.parser')
        soup = BeautifulSoup(response.text, 'lxml')

        # Находим все блоки с новостями
        # Ищем элементы с классом 'list-item' (основные новости на главной)
        news_blocks = soup.find_all('div', class_='cell-list__item m-no-image')

        print(f"📊 Найдено новостных блоков: {len(news_blocks)}")
        print("-" * 60)

        # Проходим по каждому новостному блоку
        for i, block in enumerate(news_blocks, 1):

            try:
                # Извлекаем заголовок новости
                # Ищем тег <a> с классом 'list-item__title'
                title_tag = block.find('a', class_='cell-list__item-link color-font-hover-only')

                # Проверяем, что тег найден
                if not title_tag:
                    continue

                # Получаем текст заголовка и удаляем лишние пробелы
                title = title_tag.text.strip()

                # Извлекаем ссылку на полную новость
                # Берем значение атрибута 'href'
                link = title_tag.get('href', '')

                # Если ссылка относительная (без http), добавляем базовый URL
                if link and not link.startswith('http'):
                    link = BASE_URL + link

                # Извлекаем краткое описание новости
                # Ищем тег <div> с классом 'list-item__text'
                description_tag = block.find('div', class_='list-item__text')
                description = description_tag.text.strip() if description_tag else 'Описание отсутствует'

                # Извлекаем время публикации
                # Ищем тег <div> с классом 'list-item__date'
                time_tag = block.find('div', class_='list-item__date')
                time_published = time_tag.text.strip() if time_tag else 'Время не указано'

                # Добавляем информацию о новости в список
                all_news.append({
                    'title': title,
                    'description': description,
                    'link': link,
                    'time': time_published,
                    'source': 'РИА Новости'
                })

                # Выводим информацию о текущей новости
                print(f"✅ Новость #{i}: {title[:60]}...")  # Выводим первые 60 символов заголовка

            except Exception as e:
                # Если возникла ошибка при парсинге конкретной новости, пропускаем её
                print(f"⚠️ Ошибка при парсинге новости #{i}: {e}")
                continue

        print("-" * 60)
        print(f"✅ Всего успешно спарсено новостей: {len(all_news)}")

        return all_news

    except requests.exceptions.Timeout:
        # Обработка таймаута (сервер не ответил за 10 секунд)
        print("⏱️ Ошибка: превышено время ожидания ответа от сервера")
        return all_news

    except requests.exceptions.ConnectionError:
        # Обработка ошибки подключения
        print("🔌 Ошибка: не удалось подключиться к серверу")
        return all_news

    except Exception as e:
        # Обработка любых других ошибок
        print(f"❌ Неизвестная ошибка: {e}")
        return all_news


# ============================================
# ЗАПУСК ПАРСИНГА
# ============================================


def save_news_to_excel(news_data, filename='ria_news.xlsx'):
    """
    Сохраняет новости в Excel файл с форматированием

    Аргументы:
        news_data: список словарей с новостями
        filename: имя файла для сохранения
    """

    # Проверяем, есть ли данные для сохранения
    if not news_data:
        print("⚠️ Нет данных для сохранения")
        return False

    try:
        # Создаем новый рабочий файл Excel
        wb = Workbook()
        ws = wb.active

        # Устанавливаем название листа
        ws.title = "Новости РИА"

        # Заголовки столбцов
        headers = ['№', 'Заголовок', 'Описание', 'Ссылка', 'Время публикации', 'Источник']

        # Стиль для заголовковp
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Записываем заголовки в первую строку
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Записываем данные о новостях
        for row_idx, news in enumerate(news_data, 2):  # Начинаем со второй строки
            ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Номер
            ws.cell(row=row_idx, column=2, value=news['title'])  # Заголовок
            ws.cell(row=row_idx, column=3, value=news['description'])  # Описание
            ws.cell(row=row_idx, column=4, value=news['link'])  # Ссылка
            ws.cell(row=row_idx, column=5, value=news['time'])  # Время
            ws.cell(row=row_idx, column=6, value=news['source'])  # Источник

        # Автоматическая ширина столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Устанавливаем ширину столбца (с небольшим запасом)
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        wb.save(filename)

        print(f"\n✅ Данные успешно сохранены в файл: {filename}")
        print(f"📊 Количество записей: {len(news_data)}")
        return True

    except Exception as e:
        print(f"❌ Ошибка при сохранении в Excel: {e}")
        return False


# ============================================
# ТЕСТИРОВАНИЕ СОХРАНЕНИЯ
# ============================================

if __name__ == "__main__":
    # Парсим новости
    news_data = parse_ria_news()

    # Сохраняем в Excel
    if news_data:
        save_news_to_excel(news_data, 'ria_news.xlsx')
